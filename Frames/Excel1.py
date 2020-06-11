from selenium import webdriver
import openpyxl

#driver = webdriver.Chrome()

path = "D:\Download_Python_File\Excel_Read.xlsx"

workbook = openpyxl.load_workbook(path)
#sheet = workbook.get_sheet_by_name("Sheet1")
sheet = workbook ["Sheet1"]

rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

# for r in range(1,rows+1):
#     for c in range(1,cols+1):
#         print(sheet.cell(row=r, column = c).value, end='  ')
#     print()

for r in range(1,6+1):
    for c in range(1,5):
        #sheet.cell(row=r,column = c, value = 'God')
        sheet.cell(row=r, column = c).value = 'Welcome'

workbook.save("D:\Download_Python_File\ExcelWrite1.xlsx")