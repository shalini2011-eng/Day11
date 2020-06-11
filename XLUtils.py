import openpyxl

def getRowCount(file_name, sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    Sheet = workbook[sheet_name]
    return (Sheet.max_row)

def getColumnCount(file_name, sheet_name):
    workbook = openpyxl.load_workbook(file_name)
    Sheet = workbook[sheet_name]
    return (Sheet.max_column)

def readData(file_name, sheet_name, rownum, columnno):
    workbook = openpyxl.load_workbook(file_name)
    Sheet= workbook[sheet_name]
    return Sheet.cell(row=rownum, column =columnno).value

def writeData(file_name, sheet_name, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file_name)
    Sheet = workbook[sheet_name]
    Sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file_name)